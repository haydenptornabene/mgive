# The following routine is an automated A/B Testing process. The end goal 
# is to create a tool that can be accessed from the mGive dashboard and conduct 
# specalized A/B testing.

# VERSION NUMBER: 1.0.0

# This version will be in house only and it requires the input of an .xlsx file. 
# All suppression criteria are built into this script; eventually we would like the 
# suppression criteria to be factored into the first menu of the tool once it's in the 
# dashboard. 

# Version 1.0.0 should only be used for Red Cross

# In general, this routine: 
# a1. Delete any A/B Testing Column in the Client Database 
#  1. Reads an xlsx file 
#  2. Suppresses duplicates, donors who have donated in the last 15 days, or are subscribed to Biomed
#  3. Ascrbies randomized A or B values to all members in list
#  4. Updates the Client Database with values and a new time stamp column (for filtering) 
#  5. Prints an updated Excel file for bookkeeping.     

import xlrd
import re
from xlrd import open_workbook, cellname
import csv
import string
import sys
import itertools
import openpyxl
import glob
import os
from pyexcel.cookbook import merge_all_to_a_book
import random


# For new users, run the following commands after you have installed the principle python library
# pip install xlrd
# pip install openpyxl
# pip install pyexcel
# pip install pyexcel-xlsx 



# unix code: python abtest.py mobilesubscribers.file mobiledonations.file biomed.file output.file




######## PREAMBLE ##########

# This will take the second to last argument in the command line, the input file name
subscribersin = sys.argv[-3]
donationsin = sys.argv[-2]
biomedin = sys.argv[-1]


# This will take the last argument in the command line, the output file name
#fileout = sys.argv[-1] 

# Open the desired file
booksub = open_workbook(subscribersin)
sheetsub = booksub.sheet_by_index(0)

# Open the desired file
bookusr = open_workbook(donationsin)
sheetusr = bookusr.sheet_by_index(0)

# Open the desired file
bookbio = open_workbook(biomedin)
sheetbio = bookbio.sheet_by_index(0)


#CREATE THE DICTIONARIES 
# Create a dictionary with every value from the spreadsheet from the subscriber xlsx
subkeys = [sheetsub.cell(0, col_index).value for col_index in xrange(sheetsub.ncols)]
sub_dict1 = []
for row_index in xrange(1, sheetsub.nrows):
    d = {subkeys[col_index]: sheetsub.cell(row_index, col_index).value 
         for col_index in xrange(sheetsub.ncols)}
    sub_dict1.append(d)
# Create a dictionary with every value from the spreadsheet from the donations xlsx
usrkeys = [sheetusr.cell(0, col_index).value for col_index in xrange(sheetusr.ncols)]
usr_dict1 = []
for row_index in xrange(1, sheetusr.nrows):
    d = {usrkeys[col_index]: sheetusr.cell(row_index, col_index).value 
         for col_index in xrange(sheetusr.ncols)}
    usr_dict1.append(d)
# Create a dictionary with every value from the spreadsheet from the biomed xlsx
biokeys = [sheetbio.cell(0, col_index).value for col_index in xrange(sheetbio.ncols)]
bio_dict1 = []
for row_index in xrange(1, sheetbio.nrows):
    d = {biokeys[col_index]: sheetbio.cell(row_index, col_index).value 
         for col_index in xrange(sheetbio.ncols)}
    bio_dict1.append(d)

# REMOVE DUPLICAES 
# Now that all three text files have been uploaded, we can begin to remove duplicates.
# We want to remove: 
# 1. Duplicates 2. Donors who made donations in the last 15 days 3. Biomed Subscribers 
# The sub_dict is all people subscribed in some capacity to Red Cross, the master so to speak
# We want to remove anyone from sub_dict that appears on usr_dict or bio_dict 
# (Donated in past 15 days or is a biomed subscriber)

# Isolate T-Mobile, Verizon, AT&T, and Sprint subscribers

sub_dict = []
usr_dict = []
bio_dict = []

i = 0
while i < len(sub_dict1):
    if 'Sprint' in sub_dict1[i].values():
        sub_dict.append(sub_dict1[i])
    if 'Verizon Wireless' in sub_dict1[i].values():
        sub_dict.append(sub_dict1[i])
    if 'AT&T Wireless' in sub_dict1[i].values():
        sub_dict.append(sub_dict1[i])
    if 'T-Mobile' in sub_dict1[i].values():
        sub_dict.append(sub_dict1[i])    
    i += 1

i = 0
while i < len(usr_dict1):
    if 'Sprint' in usr_dict1[i].values():
        usr_dict.append(usr_dict1[i])
    if 'Verizon Wireless' in usr_dict1[i].values():
        usr_dict.append(usr_dict1[i])
    if 'AT&T Wireless' in usr_dict1[i].values():
        usr_dict.append(usr_dict1[i])
    if 'T-Mobile' in usr_dict1[i].values():
        usr_dict.append(usr_dict1[i])    
    i += 1

i = 0
while i < len(bio_dict1):
    if 'Sprint' in bio_dict1[i].values():
        bio_dict.append(bio_dict1[i])
    if 'Verizon Wireless' in bio_dict1[i].values():
        bio_dict.append(bio_dict1[i])
    if 'AT&T Wireless' in bio_dict1[i].values():
        bio_dict.append(bio_dict1[i])
    if 'T-Mobile' in bio_dict1[i].values():
        bio_dict.append(bio_dict1[i])    
    i += 1


# Extract mobile numbers from each list
sub_mobiles = [d['Mobile Number'] for d in sub_dict]
usr_mobiles = [d['Mobile Number'] for d in usr_dict]
bio_mobiles = [d['MobileNumber'] for d in bio_dict]

# Create an intersection between sub_mobiles and usr_mobiles. 
# Create new list with intersections removed.
sub_minus_usr_mobiles = []
sub_usr_intersect = set(sub_mobiles).intersection(usr_mobiles)
for i in sub_mobiles:
	if i not in sub_usr_intersect:
		sub_minus_usr_mobiles.append(i)

# Create an intersection between sub_minus_usr_mobiles and bio_mobiles. 
# Create new  list with intersections removed. 
cleanned_mobiles = []
subusr_bio_intersect = set(sub_minus_usr_mobiles).intersection(bio_mobiles)
for i in sub_minus_usr_mobiles:
	if i not in subusr_bio_intersect:
		cleanned_mobiles.append(i)

# Remove duplicates from cleanned_mobiles
master_mobiles = []
for i in cleanned_mobiles:
	if i not in master_mobiles:
		master_mobiles.append(i)

#Initalize AB Randomizer Variables
half = len(master_mobiles)/2
oddhalf = half + 1  
A = 'A'
B = 'B'
alist = []
blist = []

# Create two lists, one of A's and one of B's that handles lists with an odd and even size length
if len(master_mobiles) % 2 == 1: # Odd Case
	alist = [A] * half
	blist = [B] * oddhalf
if len(master_mobiles) % 2 == 0: # Even Case
	alist = [A] * half
	blist = [B] * half
# Concatenate lists
ABlist = alist+blist
# Randomize list
random.shuffle(ABlist)


# Initalize A25, A10, B25, B10 Variables 
quarter = len(master_mobiles)/4
oddquarter = quarter + 1  
A25 = 'A25'
B25 = 'B25'
A10 = 'A10'
B10 = 'B10'
a25list = []
b25list = []
a10list = []
b10list = []

if len(master_mobiles) % 2 == 1: # Odd Case
    a25list = [A25] * quarter
    b25list = [B25] * oddquarter
    a10list = [A10] * oddquarter
    b10list = [B10] * oddquarter

if len(master_mobiles) % 2 == 0: # Even Case
    a25list = [A25] * quarter
    b25list = [B25] * quarter
    a10list = [A10] * quarter
    b10list = [B10] * quarter

AB2510list = a25list+b25list+a10list+b10list
random.shuffle(AB2510list)

# We create a dictionary with mobile numbers as the key and A or B as the value.
cleanned_mobiles_dictionary1 = dict(zip(master_mobiles, ABlist))

# We create a dictionary with mobile numbers as the key and A25 or B25 or A10 or B10 as the value.
cleanned_mobiles_dictionary2 = dict(zip(master_mobiles, AB2510list))

# Print dictionary to csv file to be read into SQL database. 
with open('ABTestingOutput.csv', 'wb') as csv_file:
    writer = csv.writer(csv_file)
    for key, value in cleanned_mobiles_dictionary2.items():
       writer.writerow([key, value])

print 'Your AB Test spreadsheet is ready.'

#_____________________________________________________________________________________

# We want to print out a list of SQL commands to sql executable file
# Here is the SQL command that will fill the sql executable 

# UPDATE dbo.MobileNumbers SET ABValue = "[ABValue]" WHERE MobileNumber = "[Mobile#]"
# Where [ABValue] and [Mobile#] are variables. An example is below. 

# UPDATE dbo.MobileNumbers SET ABValue = "A10" WHERE MobileNumber = "(303) 657-3331"

# We want a tool that will iterate through cleanned_mobiles_dictionary2 and populate the sql command

ListOCommands = []

# Build a loop that iterates through Cleanned dictionaries and creates a SQL command with an ABValue and the Phone Number 
for key, value in cleanned_mobiles_dictionary2.iteritems():
    Unit =  """UPDATE dbo.MobileNumbers SET ABValue=""" + "\'" + value + "\'" + " WHERE MobileNumber="  + "\'" + key + "\'"
    ListOCommands.append(Unit) 

# Removing unwanted characters from phone Numbers 
i = 0
while i < len(ListOCommands): 
    ListOCommands[i] = ListOCommands[i].replace('-', '').replace('(', '').replace(') ', '')
    i += 1

# Print the commands, row by row, in a single column, to the csv.

with open('ABTest_SQLCommands.csv', 'wb') as f:
    writer = csv.writer(f)
    for val in ListOCommands:
        writer.writerow([val])

print 'You SQL csv is also ready.'


# THIS THING IS PRETTY SLOW FOR ANYTHING BIGGER THAN 10,000
# Need to think about how to make this faster.
# But it works as is!




