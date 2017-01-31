# The following routine is meant to take an Intranet file with all Active and Inactive
# clients, containing info like CLient ID, address, email address, phone number, etc.,
# and format it in a way that is consistent with the QuickBook format for upload 
# Quickbook is the now Financial Management System 

# The final version will allow the user to drag the excel file onto a desktop app and 
# then have it spit out three corresponding spreadsheets:
# Active Clients, Inactive Clients, All (Both) Clients 
# Output Excel files will take corresponding date from input excel file

import xlrd
import re
from xlrd import open_workbook, cellname
import csv
import string
import sys
import itertools
import openpyxl
import glob
import os
from pyexcel.cookbook import merge_all_to_a_book
import random
from collections import OrderedDict
import codecs
import datetime
import math
from datetime import date, timedelta


# This is necessary to avoid the following error: 
# UnicodeEncodeError: 'ascii' codec can't encode character u'\xa0' in position 20: ordinal not in range(128)
import sys
reload(sys)
sys.setdefaultencoding("utf8")

# For new users, run the following commands after you have installed the principle python library
# pip install xlrd
# pip install openpyxl
# pip install pyexcel
# pip install pyexcel-xlsx 



# unix command: python QBFinance.py mm dd yyyy
# unix command example: python QBFinance.py 03 02 2017
							  
#______________________________________________________________________________________#


# Open the desired file
booksub = open_workbook('1.26.17ActiveCustomersIntranet.xlsx')
CompanySheet = booksub.sheet_by_index(0)
ClientSheet = booksub.sheet_by_index(1)

#CREATE THE DICTIONARIES 
# Create a dictionary with every value from the spreadsheet from the Company Profiles SpreadSheet
CompanyKeys = [CompanySheet.cell(0, col_index).value for col_index in xrange(CompanySheet.ncols)]
CompanyData = []
for row_index in xrange(1, CompanySheet.nrows):
    d = {CompanyKeys[col_index]: CompanySheet.cell(row_index, col_index).value 
         for col_index in xrange(CompanySheet.ncols)}
    CompanyData.append(d)


# Create a dictionary with every value from the spreadsheet from the Client Profiles SpreadSheet
# THIS ONE HAS RENEWAL DATE
ClientKeys = [ClientSheet.cell(0, col_index).value for col_index in xrange(ClientSheet.ncols)]
ClientData = []
for row_index in xrange(1, ClientSheet.nrows):
    d = {ClientKeys[col_index]: ClientSheet.cell(row_index, col_index).value 
         for col_index in xrange(ClientSheet.ncols)}
    ClientData.append(d)

# Client Status, when inactive, returns the string/integer '42'

# VERSION 1 - THIS VERSION ALLOWS A ONE TIME TRANSFER, IE IT IS NOT DEPENDENT ON DATE.
# VERSION 2 WILL HAVE A DATE FEATURE THAT WILL ALLOW TO FILTER BY RENEWAL DATE. 

# We need to reorder the dictionary such that the final spreadsheet matches the Quickbook format
# We will insert extra columns and Renewal Date after we reorder things. 

keyorder = ['Contact', 'Name', 'contact email', 'phone #', 'Address 2', 'Address 1', 'City', 'State', 'Zip code', 'Client ID', 'Client Status']

# Create lists Dict0-Dictn) 

AllList = [[]]*len(CompanyData)

i = 0
while i < len(CompanyData):
	AllList[i] = OrderedDict()
	for k in keyorder:
		AllList[i][k] = CompanyData[i][k]
	i += 1

# Separate into Active and Inactive 

ActiveList = []
InactiveList = []

# 0.0 is the signifier of an inactive client 
i = 0
while i < len(CompanyData):
	if 0.0 in AllList[i].values():
		InactiveList.append(AllList[i])
	else:
		ActiveList.append(AllList[i])
	i += 1

# Now we have three lists, AllList, ActiveList, and InactiveList
# Let's print them to three different CSV files

with open('AllClients.csv', 'w') as outfile:
    fp = csv.DictWriter(outfile, AllList[0].keys())
    fp.writeheader()
    fp.writerows(AllList)

with open('ActiveClients.csv', 'w') as outfile:
    fp = csv.DictWriter(outfile, AllList[0].keys())
    fp.writeheader()
    fp.writerows(ActiveList)

with open('InactiveClients.csv', 'w') as outfile:
    fp = csv.DictWriter(outfile, AllList[0].keys())
    fp.writeheader()
    fp.writerows(InactiveList)

print 'First three files have been outputted.'



#______________________________________________________________________________________

# Dates are being read in as days after 1900. 
# For example, 42913 = 6/27/2017, 42913 days after Jan 1. 1900.

# The above outputs the three files, Active, Inactive, and All. 
# The below section outputs one additional file, ActiveAfterDate.csv 

# Input Date on Command Line 
datemonth = int(sys.argv[-3])
dateday = int(sys.argv[-2])
dateyear = int(sys.argv[-1])

# Take the inputted date, subtract December 30, 1899 to account for leap year) 
# Gives date to test against
testdate = datetime.datetime(dateyear,datemonth,dateday) - datetime.datetime(1899,12,30)

# math.ceil rounds up, float("{0:.1f}".format(thing)) takes float to one decimal value
# testdate.total_seconds converts the datetime to seconds
# The constant converts from seconds to days, the desired format
floatdate = math.ceil(float("{0:.1f}".format(testdate.total_seconds()*1.15740e-5)))

ActiveListAfterDate = []
ActiveListBeforeDate = []

# We need to put the dates from ClientData into the CompanyData, but do it rigorously.

i = 0
while i < len(ClientData):
	# Find the name of a company in Company Data and return that dictionary in the variable
	match = next((l for l in CompanyData if l['Name'] == ClientData[i]['CompanyName']), None)
	# Extract the date from the ClientData List of Dictionaries 
	thisdate = ClientData[i]['RenewalDate']

	# Convert date from days since 1900 to an actual date.
	days = thisdate
	start = date(1899,12,30)      
	delta = timedelta(days)     
	newdate = start + delta  
	

	# Add the date to the dictionary entry, only if a match was found 
	# If a match is not found, the variable 'match' has data type NoneType
	if type(match) == dict:
		match['RenewalDate'] = newdate
		if thisdate >= floatdate:
			ActiveListAfterDate.append(match)
		else:
			ActiveListBeforeDate.append(match)
	i += 1


# Reorganize the list so it has the correct format. 
keyorderRD = ['Contact', 'Name', 'contact email', 'phone #', 'Address 2', 'Address 1', 'City', 'State', 'Zip code', 'Client ID', 'Client Status', 'RenewalDate']
ALADS = [[]]*len(ActiveListAfterDate)

i = 0
while i < len(ActiveListAfterDate):
	ALADS[i] = OrderedDict()
	for k in keyorderRD:
		ALADS[i][k] = ActiveListAfterDate[i][k]
	i += 1


# Remove Duplicates from ALADS
ALAD = []
for x in ALADS:
    if x not in ALAD:
        ALAD.append(x)

# Output list to ActiveAfterDate.csv

with open('ActiveAfterDate.csv', 'w') as outfile:
    fp = csv.DictWriter(outfile, ALAD[0].keys())
    fp.writeheader()
    fp.writerows(ALAD)

print 'Your last, date specific, file has been outputted.'
print 'Have a wonderful day, Chris :)'









